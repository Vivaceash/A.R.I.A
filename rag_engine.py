import os
import re
import math
import zipfile
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from pypdf import PdfReader
import openpyxl

# --- PARSERS DE DOCUMENTOS ---

def parse_text_file(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception as e:
        print(f"Error parseando archivo de texto {filepath}: {e}")
        return ""

def parse_docx_file(filepath):
    try:
        with zipfile.ZipFile(filepath) as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            texts = []
            for elem in root.iter():
                if elem.tag.endswith('t'):
                    if elem.text:
                        texts.append(elem.text)
            return " ".join(texts)
    except Exception as e:
        print(f"Error parseando docx {filepath}: {e}")
        return ""

def parse_pdf_file(filepath):
    try:
        reader = PdfReader(filepath)
        texts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
        return "\n".join(texts)
    except Exception as e:
        print(f"Error parseando PDF {filepath}: {e}")
        return ""

def parse_xlsx_file(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        texts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            texts.append(f"Hoja: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_str = " | ".join(str(cell) for cell in row if cell is not None)
                if row_str.strip():
                    texts.append(row_str)
        return "\n".join(texts)
    except Exception as e:
        print(f"Error parseando XLSX {filepath}: {e}")
        return ""

class HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.text_parts = []
    def handle_data(self, data):
        cleaned = data.strip()
        if cleaned:
            self.text_parts.append(cleaned)
    def get_text(self):
        return " ".join(self.text_parts)

def parse_xls_html_file(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        if "<html>" in content.lower() or "<body>" in content.lower() or "<table" in content.lower():
            parser = HTMLTextExtractor()
            parser.feed(content)
            return parser.get_text()
        else:
            return content
    except Exception as e:
        print(f"Error parseando XLS HTML {filepath}: {e}")
        return ""

def extract_file_content(filepath):
    _, ext = os.path.splitext(filepath.lower())
    if ext in ['.txt', '.py', '.json', '.csv', '.md', '.log', '.js', '.css', '.html']:
        return parse_text_file(filepath)
    elif ext == '.docx':
        return parse_docx_file(filepath)
    elif ext == '.pdf':
        return parse_pdf_file(filepath)
    elif ext == '.xlsx':
        return parse_xlsx_file(filepath)
    elif ext == '.xls':
        return parse_xls_html_file(filepath)
    else:
        return ""

# --- CHUNKING ---

def chunk_text(text, filename, max_chars=600, overlap=150):
    text = re.sub(r'\s+', ' ', text).strip()
    chunks = []
    if not text:
        return chunks
        
    if len(text) <= max_chars:
        chunks.append({"text": text, "filename": filename})
        return chunks
        
    start = 0
    while start < len(text):
        end = min(start + max_chars, len(text))
        if end < len(text):
            split_idx = text.rfind('.', start + max_chars // 2, end)
            if split_idx == -1:
                split_idx = text.rfind(' ', start + max_chars // 2, end)
            if split_idx != -1:
                end = split_idx + 1
        chunks.append({"text": text[start:end].strip(), "filename": filename})
        start = end - overlap
        if start >= len(text) - overlap:
            break
    return chunks

def get_directory_chunks(directory_path):
    all_chunks = []
    if not os.path.exists(directory_path):
        return all_chunks
        
    for filename in os.listdir(directory_path):
        filepath = os.path.join(directory_path, filename)
        if os.path.isfile(filepath):
            if filename.startswith('~') or filename.endswith('~') or filename.endswith('#') or filename.endswith('.tmp') or filename.startswith('.'):
                continue
            content = extract_file_content(filepath)
            if content.strip():
                file_chunks = chunk_text(content, filename)
                all_chunks.extend(file_chunks)
    return all_chunks

# --- RETRIEVAL ---

def tokenize(text):
    return re.findall(r'\w+', text.lower())

def retrieve_context(query, chunks, top_k=4):
    query_tokens = tokenize(query)
    if not query_tokens or not chunks:
        return "No hay información disponible en los documentos locales en este momento."
        
    df = {}
    for chunk in chunks:
        tokens = set(tokenize(chunk["text"]))
        for t in tokens:
            df[t] = df.get(t, 0) + 1
            
    scored_chunks = []
    for chunk in chunks:
        chunk_text_lower = chunk["text"].lower()
        chunk_tokens = tokenize(chunk_text_lower)
        chunk_tokens_set = set(chunk_tokens)
        
        score = 0.0
        for token in query_tokens:
            if token in chunk_tokens_set:
                tf = chunk_tokens.count(token)
                idf = math.log(len(chunks) / (df.get(token, 0) + 0.5)) + 1
                score += tf * idf
                
        if score > 0:
            scored_chunks.append((score, chunk))
            
    scored_chunks.sort(key=lambda x: x[0], reverse=True)
    
    selected = scored_chunks[:top_k]
    if not selected:
        return "No se encontraron fragmentos relevantes en los documentos locales para esta consulta."
        
    context_parts = []
    for idx, (score, chunk) in enumerate(selected):
        context_parts.append(f"--- FRAGMENTO {idx+1} (Archivo: {chunk['filename']}) ---\n{chunk['text']}")
        
    return "\n\n".join(context_parts)
